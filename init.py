from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from dotenv import load_dotenv

import time
import json
import shutil


load_dotenv()
chrome_options = webdriver.ChromeOptions()
settings = {
       "recentDestinations": [{
            "id": "Save as PDF",
            "origin": "local",
            "account": "",
        }],
        "selectedDestinationId": "Save as PDF",
        "version": 2
}
prefs = {'printing.print_preview_sticky_settings.appState': json.dumps(settings), 
           'savefile.default_directory': '/Users/user/Projects/serasa-consulting/consultas'
        }
chrome_options.add_experimental_option('prefs', prefs)
chrome_options.add_argument('--kiosk-printing')
CHROMEDRIVER_PATH = './chromedriver'
driver = webdriver.Chrome(chrome_options=chrome_options, executable_path=CHROMEDRIVER_PATH)

driver.get("https://www.serasaexperian.com.br/")
## Inital form login
WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="header-main"]/div[2]/nav/div[2]/button[2]'))).click()
time.sleep(1)
WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="username"]'))).send_keys(os.getenv('USERNAME'))
time.sleep(1)
WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="password"]'))).send_keys(os.getenv('PASSWORD'))
WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="form-login"]/div/div/div[2]/div/form/div[5]/div/button'))).click()

## Page after login - User Actions
time.sleep(3)
driver.refresh()
WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="card-5f21f84fcb8be941bb2470e3"]/div[1]/div[2]/button[2]'))).click()
## Open XLSX
wb = load_workbook("cpf.xlsx", data_only=True)
first_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name(first_sheet)

for row in range(2,worksheet.max_row+1):
    for column in "B":
        cell_name = "{}{}".format(column, row)
        cpf = worksheet[cell_name].value

        if cpf != None and worksheet["{}{}".format("D", row)].value == None:
            time.sleep(2)
            print("Iniciando consulta para cpf: ", cpf)
            WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="tipoDocumentoCpf"]'))).click()
            time.sleep(2)
            WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="cpfCnpjId"]'))).send_keys(cpf)
            WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="Link"]'))).click()

            time.sleep(2)
            text = "O documento consultado tem participação em empresa(s)."
            if str(text) in driver.page_source:
                WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="selecaoParticipacaoSocietaria:1"]'))).click()
                WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="table-part-societaria"]/tbody/tr[2]/td/a'))).click()
            
            ## Loading final consult
            time.sleep(3)
            debtor = False
            for i in range(4,12):
                elementPath = '//*[@id="formResultado:tbl"]/table[2]/tbody/tr[{}]/td[4]'.format(i)
                value = WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, elementPath))).text
                if value != '-' and value != '0':
                    debtor = True
            
            if debtor == True:
                worksheet["{}{}".format("D", row)].value = "Restrição"
            else:
                worksheet["{}{}".format("D", row)].value = "NÃO"
                
            print("Consulta finalizada")

            driver.execute_script('window.print();')
            shutil.move('./consultas/Credit Report, credit score and credit check from Serasa Experian.pdf', "./consultas/{}.pdf".format(cpf))
            wb.save('cpf.xlsx')
            time.sleep(2)
            WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="formRes"]/div/a[1]'))).click()
                

driver.quit()